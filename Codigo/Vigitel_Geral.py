import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles.numbers import BUILTIN_FORMATS
from Metodos.metodo_geral import Automacao_Vigitel_Geral
from io import BytesIO
import streamlit as st
from datetime import date


# Função para salvar as tabelas em um único Excel de aba única e com formatação
def salvar_excel_com_formatacao(bd_todas_cidades):
    output = BytesIO()
    #=== Salvar em uma planilha em excel ===#
    # Crie uma nova planilha Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio_Vigitel"

    # Remover as linhas de grade do Excel
    ws.sheet_view.showGridLines = False

    # Define o estilo de preenchimento para o fundo do cabeçalho
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Define o alinhamento centralizado com quebra de texto
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Define o alinhamento centralizado
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Define o tamanho da fonte para todas as células
    font_size_9 = Font(size=9)

    # Define a fonte em negrito
    bold_font = Font(bold=True, size=9)

    # Define a fonte azul em negrito
    blue_bold_font = Font(color="0000FF", bold=True, size=9)

    # Loop pelos DataFrames e escrevendo na planilha
    row_offset = 1  # Inicializa a contagem de linhas na planilha
    for i, df in enumerate(bd_todas_cidades):
        # Converter o DataFrame para linhas que o openpyxl pode usar
        rows = dataframe_to_rows(df, index=False, header=True)
        last_row_start = ws.max_row + 1  # Marca a posição da primeira linha do DataFrame atual
        
        # Escreve os dados no Excel
        for j, row in enumerate(rows):
            ws.append(row)

            # Itera sobre todas as colunas e aplica centralização
            for col in range(1, len(row) + 1):  # Itera sobre as colunas
                cell = ws.cell(row=row_offset + j, column=col)
                cell.alignment = center_alignment  # Aplica alinhamento centralizado
                cell.font = font_size_9 # Aplica o tamanho da fonte de 9 para todas as células

                # Estiliza apenas o cabeçalho
                if (j == 0) | ((j == 1) and (i == (len(bd_todas_cidades) - 1))):
                    cell.font = bold_font
                if (j == 1) and (i != (len(bd_todas_cidades) - 1)):  # Cabeçalhos do multiíndice
                    cell.fill = header_fill
                    cell.font = bold_font
                    cell.alignment = header_alignment # Estiliza o cabeçalho com quebra de texto
           
            # Estiliza a coluna "20" (quarta coluna) em azul e negrito
            if j > 1:  # Dados (exclui os cabeçalhos)
                cell = ws.cell(row=row_offset + j, column=5)  # Quinta coluna (índice 5 no Excel)
                cell.font = blue_bold_font

            # Formata como porcentagem colunas de 24 a 28
            if (j > 1) | ((j == 1) and (i == (len(bd_todas_cidades) - 1))):  # Dados (exclui os cabeçalhos)
                for col in range(24, 29):  # Colunas 24 a 28
                    cell = ws.cell(row=row_offset + j, column=col)
                    if isinstance(cell.value, (int, float)):  # Verifica se é número
                        cell.number_format = '0.0%'  # Formato de porcentagem com uma casa decimal
                        cell.value = float(cell.value)  # Converte o valor para floats
      
        # Aplica o estilo de negrito à última linha do DataFrame
        last_row = ws.max_row  # Identifica a última linha escrita
        for col in range(1, df.shape[1] + 1):  # Itera sobre todas as colunas do DataFrame
            cell = ws.cell(row=last_row, column=col)
            cell.font = bold_font
            cell = ws.cell(row=last_row-1, column=col)
            cell.font = bold_font
            if col == 5:
                cell.font = blue_bold_font
        
        # Adicionar 1 linha em branco entre os DataFrames, exceto no último
        if i < (len(bd_todas_cidades) - 1):
            row_offset = ws.max_row + 2  # Ajusta o offset para o próximo DataFrame
            for _ in range(1):  # Adiciona linhas vazias
                ws.append([])
        else:
            row_offset = ws.max_row + 1
    
    # Salvar o Workbook no buffer
    wb.save(output)
    return output.getvalue()




# Configurações da página
st.set_page_config(layout="centered")  # "wide"

#=== Título ===#
st.title("Automatização Vigitel")
st.write("Faça o upload dos relatórios para realizar o tratamento.")

# Upload da planilha
with st.form(key='sheet_name_data'):
    nome_relatorio = st.text_input(label="Insira o nome do relatório (texto na célula da planilha que se encontra na primeira linha e primeira coluna).")
    nome_sheet_DATA_FIXO = st.text_input(label="Insira o nome da sheet (aba) no qual contém o relatório do FIXO no formato xlsx com os dados Vigitel.")
    REPLICAS_FIXO = st.text_input(label="""Informe a quantidade de Réplicas do relatório do FIXO que terão para cada cidade em ordem conforme se encontra no excel. 
                                      Os valores deverão ser informados separados por vírgula e espaço (, ).
                                      \nExemplo: '10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 20, 20, 20, 20, 30, 30, 30, 30'""")


    nome_sheet_DATA_CEL = st.text_input(label="Insira o nome da sheet (aba) no qual contém o relatório do CELULAR no formato xlsx com os dados Vigitel.")
    REPLICAS_CEL = st.text_input(label="""Informe a quantidade de Réplicas do relatório do CELULAR que terão para cada cidade em ordem conforme se encontra no excel. 
                                      Os valores deverão ser informados separados por vírgula e espaço (, ).
                                      \nExemplo: '10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 20, 20, 20, 20, 30, 30, 30, 30'""")
    input_buttom_submit_DATA = st.form_submit_button("Enviar")


st.session_state.nome_relatorio = nome_relatorio

st.session_state.nome_sheet_DATA_FIXO = nome_sheet_DATA_FIXO
st.session_state.REPLICAS_FIXO = REPLICAS_FIXO

st.session_state.nome_sheet_DATA_CEL = nome_sheet_DATA_CEL
st.session_state.REPLICAS_CEL = REPLICAS_CEL

data_fixo = st.file_uploader("Selecione o relatório do FIXO no formato xlsx com os dados Vigitel.", type=["xlsx"])
data_cel = st.file_uploader("Selecione o relatório do CELULAR no formato xlsx com os dados Vigitel.", type=["xlsx"])

if data_fixo and data_cel:
    st.write("Planilhas carregadas com sucesso!")
    nome_relatorio = st.session_state.nome_relatorio
    nome_sheet_DATA_FIXO = st.session_state.nome_sheet_DATA_FIXO
    REPLICAS_FIXO = st.session_state.REPLICAS_FIXO
    REPLICAS_FIXO = REPLICAS_FIXO.split(', ')
    data_fixo = pd.read_excel(data_fixo, sheet_name=nome_sheet_DATA_FIXO)
    data_fixo = data_fixo.dropna(axis=0, subset=nome_relatorio)

    nome_sheet_DATA_CEL = st.session_state.nome_sheet_DATA_CEL
    REPLICAS_CEL = st.session_state.REPLICAS_CEL
    REPLICAS_CEL = REPLICAS_CEL.split(', ')
    data_cel = pd.read_excel(data_cel, sheet_name=nome_sheet_DATA_CEL)
    data_cel = data_cel.dropna(axis=0, subset=nome_relatorio)
    
    # Botão para processar os dados
    if st.button("Processar Dados"):
        # Processar os dados e obter as tabelas
        bd_todas_cidades = Automacao_Vigitel_Geral(REPLICAS_FIXO, data_fixo, REPLICAS_CEL, data_cel)
        
        # Salvar em Excel com formatação
        excel_data = salvar_excel_com_formatacao(bd_todas_cidades=bd_todas_cidades)

        # Obtendo a data atual
        data_atual = date.today()
        # Formatando a data no formato DD/MM/YYYY
        data_formatada = data_atual.strftime("%d-%m-%Y")
        # print("Data formatada:", data_formatada)
        
        # Link para download
        st.download_button(
            label="Relatorio Vigitel GERAL",
            data=excel_data,
            file_name=f"Relatorio Vigitel GERAL_{data_formatada}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )