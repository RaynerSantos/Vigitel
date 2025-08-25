# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.styles import PatternFill


# #=== Salvar as tabelas no Excel ===#
# # Crie uma nova planilha Excel
# wb = Workbook()
# ws = wb.active
# ws.title = "Cidades"

# # Lista de DataFrames (exemplo com dataframes fictícios)
# bd_todas_cidades = []

# # Exemplo de DataFrames com MultiIndex nos cabeçalhos
# df1 = pd.DataFrame([[1, 2], [3, 4]], columns=pd.MultiIndex.from_product([['Cidade1'], ['Col1', 'Col2']]))
# df2 = pd.DataFrame([[5, 6], [7, 8]], columns=pd.MultiIndex.from_product([['Cidade2'], ['Col1', 'Col2']]))
# bd_todas_cidades.extend([df1, df2])

# # Define o estilo de preenchimento para o fundo do cabeçalho
# header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# # Loop pelos DataFrames e escrevendo na planilha
# row_offset = 1  # Inicializa a contagem de linhas na planilha
# for i, df in enumerate(bd_todas_cidades):
#     # Converter o DataFrame para linhas que o openpyxl pode usar
#     rows = dataframe_to_rows(df, index=False, header=True)
    
#     # Escreve os dados no Excel
#     for j, row in enumerate(rows):
#         ws.append(row)
        
#         # Estiliza apenas o cabeçalho
#         if j == 1:  # Cabeçalho da tabela (primeira linha do DataFrame)
#             for col in range(1, len(row) + 1):  # Itera sobre as colunas
#                 ws.cell(row=row_offset + j, column=col).fill = header_fill
    
#     # Adicionar 3 linhas em branco entre os DataFrames, exceto no último
#     if i < len(bd_todas_cidades) - 1:
#         row_offset = ws.max_row + 4  # Ajusta o offset para o próximo DataFrame
#         for _ in range(1):  # Adiciona linhas vazias
#             ws.append([])
#     else:
#         row_offset = ws.max_row + 1

# # Salvar o arquivo Excel
# wb.save("dataframes_cidades_estilizado.xlsx")






########################################################################################################################






import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.numbers import BUILTIN_FORMATS

# TOTAL = 400
TOTAL = '0, 400, 400, 0, 400, 400, 0, 400, 400, 400, 400, 0, 0, 0, 400, 0, 0, 0, 0, 400, 0, 400, 400, 0, 400, 0, 0'
TOTAL = TOTAL.split(', ')
REPLICAS = '10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10' # Fixo
# REPLICAS = '30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30' # Celular
REPLICAS = REPLICAS.split(', ')

COLUNAS = ['Réplica', 'TOTAL', 'Total Elegíveis', '20', '5', '6', '4', '44', '66', '7', '8', '9', '10', '55', '88 Eleg', 
           'Total Não Eleg', '1', '2', '3', '7 N.E', '22', 'Out. Cid. 33', '88 Não Eleg', 
           'Tx. Elegível', 'Tx. Sucesso', 'Recusa Total', 'Recusa Agenda', 'Recusa Entrev.', 'Virgens']

# bd = pd.read_excel("relatorio_elegiveis_fixo_teste.xlsx")
bd = pd.read_excel("C:\PROJETOS\Vigitel\TESTE 13-01\Relatório de Elegíveis_Fixo_27_EDITADO.xlsx")
print(f'BD: {bd}\n')


COLUNAS = ['Réplica', 'TOTAL', 'Total Elegíveis', '20', '5', '6', '4', '44', '66', '7', '8', '9', '10', '55', '88 Eleg', 
        'Total Não Eleg', '1', '2', '3', '7 N.E', '22', 'Out. Cid. 33', '88 Não Eleg', 
        'Tx. Elegível', 'Tx. Sucesso', 'Recusa Total', 'Recusa Agenda', 'Recusa Entrev.', 'Virgens']

def buscar_cidades(SOMA_LINHAS=0, SOMA_FIXA=3, REPLICAS=[], bd=pd.DataFrame()):
    lista_cidades = []
    for i, rep in enumerate(REPLICAS):
        if i == 0:
            print(bd.columns[0])
            lista_cidades.append(bd.columns[0])
        elif i == 1:
            print(bd.iloc[SOMA_LINHAS + int(rep) + SOMA_FIXA, 0])
            lista_cidades.append(bd.iloc[SOMA_LINHAS + int(rep) + SOMA_FIXA, 0])
            SOMA_LINHAS += int(rep) + SOMA_FIXA 
        else:
            print(bd.iloc[SOMA_LINHAS + int(rep) + SOMA_FIXA + 1, 0])
            lista_cidades.append(bd.iloc[SOMA_LINHAS + int(rep) + SOMA_FIXA + 1, 0])
            SOMA_LINHAS += int(rep) + SOMA_FIXA + 1
    return lista_cidades

CIDADES = buscar_cidades(SOMA_LINHAS=0, SOMA_FIXA=3, REPLICAS=REPLICAS, bd=bd)


bd_todas_cidades = []
CONTAGEM_LINHAS = 4
SOMA_LINHAS = 0

df_total_geral = pd.DataFrame(columns=[
'Réplica', 'TOTAL', 'Total Elegíveis', '20', '5', '6', '4', '44', '66', '7', '8', '9', '10', 
'88 Eleg', 'Total Não Eleg', '1', '2', '3', '7 N.E', '22', 'Out. Cid. 33', '88 Não Eleg', 
'Tx. Elegível', 'Tx. Sucesso', 'Recusa Total', 'Recusa Agenda', 'Recusa Entrev.', 'Virgens'
])

for j, cidade in enumerate(CIDADES):
    print(f'\nVERIFICAR O TOTAL:\t{int(TOTAL[j])}')
    if int(TOTAL[j]) == 0:
        #===== Selecionando o dataframe para cada cidade =====#
        if j == 0:
            tabela = bd.loc[(2 + SOMA_LINHAS):(int(REPLICAS[j]) + 2), :]
        else:
            tabela = bd.loc[(2 + SOMA_LINHAS):(SOMA_LINHAS + int(REPLICAS[j]) + 2), :]
        SOMA_LINHAS += int(REPLICAS[j]) + CONTAGEM_LINHAS
        # aracaju.columns = bd_fixo.loc[0, :]
        tabela.columns = COLUNAS
        tabela = tabela[['Réplica', 'TOTAL', 'Total Elegíveis', '20', '5', '6', '4', '44', '66', '7', '8', '9', '10', '88 Eleg', 
        'Total Não Eleg', '1', '2', '3', '7 N.E', '22', 'Out. Cid. 33', '88 Não Eleg', 
        'Tx. Elegível', 'Tx. Sucesso', 'Recusa Total', 'Recusa Agenda', 'Recusa Entrev.', 'Virgens']]
        tabela.reset_index(inplace=True)
        tabela = tabela.iloc[:, 1:]
        # print(f'TABELA: {tabela}')

    else:
        #===== Selecionando o dataframe para cada cidade =====#
        if j == 0:
            tabela = bd.loc[(2 + SOMA_LINHAS):(int(REPLICAS[j]) + 2), :]
        else:
            tabela = bd.loc[(2 + SOMA_LINHAS):(SOMA_LINHAS + int(REPLICAS[j]) + 2), :]
        SOMA_LINHAS += int(REPLICAS[j]) + CONTAGEM_LINHAS
        # aracaju.columns = bd_fixo.loc[0, :]
        tabela.columns = COLUNAS
        tabela = tabela[['Réplica', 'TOTAL', 'Total Elegíveis', '20', '5', '6', '4', '44', '66', '7', '8', '9', '10', '88 Eleg', 
        'Total Não Eleg', '1', '2', '3', '7 N.E', '22', 'Out. Cid. 33', '88 Não Eleg', 
        'Tx. Elegível', 'Tx. Sucesso', 'Recusa Total', 'Recusa Agenda', 'Recusa Entrev.', 'Virgens']]
        tabela.reset_index(inplace=True)
        tabela = tabela.iloc[:, 1:]
        
        #===== Tratando 7 N.E | TOTAL | Total Não Eleg =====#
        for k in range(2):
            for i in range(len(tabela["7 N.E"])):

                # Tratamento da coluna "Total Elegíveis"
                soma_eleg = tabela.loc[i, ['20', '5', '6', '4', '44', '66', '7', '8', '9', '10', '88 Eleg']].sum()
                tabela.loc[i, "Total Elegíveis"] = soma_eleg

                # Tratando o "7 N.E"
                if int(tabela.loc[i, "TOTAL"]) != int(TOTAL[j]):
                    valor_a_somar = int(TOTAL[j]) - tabela.loc[i, "TOTAL"]
                    tabela.loc[i, "7 N.E"] = tabela.loc[i, "7 N.E"] + valor_a_somar
                
                # Tratando o "Total Não Eleg"
                soma_nao_eleg = tabela.loc[i, ['1','2','3','7 N.E','22','Out. Cid. 33','88 Não Eleg']].sum()
                tabela.loc[i, "Total Não Eleg"] = soma_nao_eleg

                # Tratando o "TOTAL"
                soma_TOTAL = tabela.loc[i, ["Total Elegíveis", "Total Não Eleg", "Virgens"]].sum()
                tabela.loc[i, "TOTAL"] = soma_TOTAL
        
        #===== Tx. Elegível =====#
        for i in range(len(tabela["Tx. Elegível"])):
            denominador = ( tabela.loc[i, "TOTAL"] - tabela.loc[i, "Virgens"] )
            numerador = tabela.loc[i, "Total Elegíveis"]
            if denominador == 0:
                tabela.loc[i, "Tx. Elegível"] = 0
            else:
                tabela.loc[i, "Tx. Elegível"] = numerador / denominador

        #===== Tx. Sucesso =====#
        for i in range(len(tabela["Tx. Sucesso"])):
            denominador = tabela.loc[i, "Total Elegíveis"]
            numerador = tabela.loc[i, "20"]
            if denominador == 0:
                tabela.loc[i, "Tx. Sucesso"] = 0
            else:
                tabela.loc[i, "Tx. Sucesso"] = numerador / denominador
        
        #===== Recusa Agenda | Recusa Entrev. | Recusa Total =====#
        for i in range(len(tabela["Recusa Agenda"])):
            denominador = tabela.loc[i, "Total Elegíveis"]
            numerador_rec_agenda = tabela.loc[i, "4"]
            numerador_rec_entrev = tabela.loc[i, "44"]
            if denominador == 0:
                tabela.loc[i, "Recusa Agenda"] = 0
                tabela.loc[i, "Recusa Entrev."] = 0
            else:
                tabela.loc[i, "Recusa Agenda"] = numerador_rec_agenda / denominador
                tabela.loc[i, "Recusa Entrev."] = numerador_rec_entrev / denominador
            tabela.loc[i, "Recusa Total"] = (tabela.loc[i, "Recusa Agenda"] + tabela.loc[i, "Recusa Entrev."])
        
        #===== Tratar a última linha de sub Total e rodar todas as formulas para calcular as taxas novamente =====#
        tabela.iloc[ (int(REPLICAS[j])), : ] = tabela.iloc[ :(int(REPLICAS[j])), : ].sum()

        for i in range(len(tabela["Tx. Elegível"])):
            denominador = ( tabela.loc[i, "TOTAL"] - tabela.loc[i, "Virgens"] )
            numerador = tabela.loc[i, "Total Elegíveis"]
            if denominador == 0:
                tabela.loc[i, "Tx. Elegível"] = 0
            else:
                tabela.loc[i, "Tx. Elegível"] = numerador / denominador


        for i in range(len(tabela["Tx. Sucesso"])):
            denominador = tabela.loc[i, "Total Elegíveis"]
            numerador = tabela.loc[i, "20"]
            if denominador == 0:
                tabela.loc[i, "Tx. Sucesso"] = 0
            else:
                tabela.loc[i, "Tx. Sucesso"] = numerador / denominador


        for i in range(len(tabela["Recusa Agenda"])):
            denominador = tabela.loc[i, "Total Elegíveis"]
            numerador_rec_agenda = tabela.loc[i, "4"]
            numerador_rec_entrev = tabela.loc[i, "44"]
            if denominador == 0:
                tabela.loc[i, "Recusa Agenda"] = 0
                tabela.loc[i, "Recusa Entrev."] = 0
            else:
                tabela.loc[i, "Recusa Agenda"] = numerador_rec_agenda / denominador
                tabela.loc[i, "Recusa Entrev."] = numerador_rec_entrev / denominador
            tabela.loc[i, "Recusa Total"] = (tabela.loc[i, "Recusa Agenda"] + tabela.loc[i, "Recusa Entrev."])
    
    tabela.iloc[(int(REPLICAS[j])), 0] = "Sub Total:"

    # Copiar a linha 10 do DataFrame original
    linha_subtotal = tabela.iloc[int(REPLICAS[j]), :].to_frame().T

    # Concatenar a linha com o DataFrame vazio
    df_total_geral = pd.concat([df_total_geral, linha_subtotal], ignore_index=True)

    replica_ativa = []
    for i in range(len(tabela["Tx. Elegível"])):
        if tabela.loc[i, "Tx. Elegível"] > 0:
            replica_ativa.append("SIM")
        else:
            replica_ativa.append("NÃO")
    tabela["Réplica Ativa"] = replica_ativa
    tabela = tabela[['Réplica', 'Réplica Ativa', 'TOTAL', 'Total Elegíveis', '20', '5', '6', '4', '44', '66', '7', '8', '9', '10', '88 Eleg', 
        'Total Não Eleg', '1', '2', '3', '7 N.E', '22', 'Out. Cid. 33', '88 Não Eleg', 
        'Tx. Elegível', 'Tx. Sucesso', 'Recusa Total', 'Recusa Agenda', 'Recusa Entrev.', 'Virgens']]
    tabela.loc[int(REPLICAS[j]), 'Réplica Ativa'] = ''

    tabela.loc[:, 'Total Tentativas'] = ''
    
    tabela.columns = pd.MultiIndex.from_product([[cidade], tabela.columns])
    
    bd_todas_cidades.append(tabela)
    
    print(bd_todas_cidades[j])


#===== Última tabela =====#
df_total_geral = df_total_geral.sum().to_frame().T
df_total_geral['Réplica'] = 'Total geral:'
df_total_geral.insert(1, 'Réplica Ativa', '')
#===== Tx. Elegível =====#
denominador = ( df_total_geral.loc[0, "TOTAL"] - df_total_geral.loc[0, "Virgens"] )
numerador = df_total_geral.loc[0, "Total Elegíveis"]
if denominador == 0:
    df_total_geral.loc[0, "Tx. Elegível"] = 0
else:
    df_total_geral.loc[0, "Tx. Elegível"] = numerador / denominador

#===== Tx. Sucesso =====#
denominador = df_total_geral.loc[0, "Total Elegíveis"]
numerador = df_total_geral.loc[0, "20"]
if denominador == 0:
    df_total_geral.loc[0, "Tx. Sucesso"] = 0
else:
    df_total_geral.loc[0, "Tx. Sucesso"] = numerador / denominador


#===== Recusa Agenda | Recusa Entrev. | Recusa Total =====#
denominador = df_total_geral.loc[0, "Total Elegíveis"]
numerador_rec_agenda = df_total_geral.loc[0, "4"]
numerador_rec_entrev = df_total_geral.loc[0, "44"]
if denominador == 0:
    df_total_geral.loc[0, "Recusa Agenda"] = 0
    df_total_geral.loc[0, "Recusa Entrev."] = 0
else:
    df_total_geral.loc[0, "Recusa Agenda"] = numerador_rec_agenda / denominador
    df_total_geral.loc[0, "Recusa Entrev."] = numerador_rec_entrev / denominador
df_total_geral.loc[0, "Recusa Total"] = (df_total_geral.loc[0, "Recusa Agenda"] + df_total_geral.loc[0, "Recusa Entrev."])

nova_coluna_df_total_geral = [''] * 29
df_total_geral.columns = nova_coluna_df_total_geral

bd_todas_cidades.append(df_total_geral)