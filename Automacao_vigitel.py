import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.numbers import BUILTIN_FORMATS

# CIDADES = 'ARACAJU, BELEM, BELO HORIZONTE, BOA VISTA, BRASILIA, CAMPO GRANDE, CUIABA, CURITIBA, FLORIANOPOLIS, FORTALEZA, GOIANIA, JOAO PESSOA, MACAPA, MACEIO, MANAUS, NATAL, PALMAS, PORTO ALEGRE, PORTO VELHO, RECIFE, RIO BRANCO, RIO DE JANEIRO, SALVADOR, SAO LUIS, SAO PAULO, TERESINA, VITORIA'
# CIDADES = CIDADES.split(', ')
TOTAL = 400
REPLICAS = '10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10' # Fixo
# REPLICAS = '30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30, 30' # Celular
REPLICAS = REPLICAS.split(', ')

COLUNAS = ['Réplica', 'TOTAL', 'Total Elegíveis', '20', '5', '6', '4', '44', '66', '7', '8', '9', '10', '55', '88 Eleg', 
           'Total Não Eleg', '1', '2', '3', '7 N.E', '22', 'Out. Cid. 33', '88 Não Eleg', 
           'Tx. Elegível', 'Tx. Sucesso', 'Recusa Total', 'Recusa Agenda', 'Recusa Entrev.', 'Virgens']

bd = pd.read_excel("relatorio_elegiveis_fixo_teste.xlsx")
bd


def buscar_cidades(SOMA_LINHAS=0, SOMA_FIXA=3, REPLICAS=[], bd=pd.DataFrame()):
    lista_cidades = []
    for i, rep in enumerate(REPLICAS):
        if i == 0:
            print(bd.columns[0])
            lista_cidades.append(bd.columns[0])
        elif i == 1:
            print(bd.iloc[SOMA_LINHAS + int(rep) + SOMA_FIXA, 0])
            lista_cidades.append(bd.iloc[SOMA_LINHAS + int(rep) + SOMA_FIXA, 0])
            SOMA_LINHAS += int(rep) + SOMA_FIXA 
        else:
            print(bd.iloc[SOMA_LINHAS + int(rep) + SOMA_FIXA + 1, 0])
            lista_cidades.append(bd.iloc[SOMA_LINHAS + int(rep) + SOMA_FIXA + 1, 0])
            SOMA_LINHAS += int(rep) + SOMA_FIXA + 1
    return lista_cidades

CIDADES = buscar_cidades(SOMA_LINHAS=0, SOMA_FIXA=3, REPLICAS=REPLICAS, bd=bd)


bd_todas_cidades = []
CONTAGEM_LINHAS = 4
SOMA_LINHAS = 0
QTD_LINHAS_TABELA = int(REPLICAS[0])

for j, cidade in enumerate(CIDADES):
    #===== Selecionando o dataframe para cada cidade =====#
    if j == 0:
        tabela = bd.loc[(2 + SOMA_LINHAS):(int(REPLICAS[j]) + 2), :]
    else:
        tabela = bd.loc[(2 + SOMA_LINHAS):(SOMA_LINHAS + int(REPLICAS[j]) + 2), :]
    SOMA_LINHAS += int(REPLICAS[j]) + CONTAGEM_LINHAS
    # aracaju.columns = bd_fixo.loc[0, :]
    tabela.columns = COLUNAS
    tabela.reset_index(inplace=True)
    tabela = tabela.iloc[:, 1:]
    
    #===== Tratando 7 N.E | TOTAL | Total Não Eleg =====#
    for k in range(2):
        for i in range(len(tabela["7 N.E"])):
            # Tratando o "7 N.E"
            if tabela.loc[i, "TOTAL"] != TOTAL:
                valor_a_somar = TOTAL - tabela.loc[i, "TOTAL"]
                tabela.loc[i, "7 N.E"] = tabela.loc[i, "7 N.E"] + valor_a_somar
            
            # Tratando o "Total Não Eleg"
            soma_nao_eleg = tabela.loc[i, ['1','2','3','7 N.E','22','Out. Cid. 33','88 Não Eleg']].sum()
            tabela.loc[i, "Total Não Eleg"] = soma_nao_eleg

            # Tratando o "TOTAL"
            soma_TOTAL = tabela.loc[i, ["Total Elegíveis", "Total Não Eleg", "Virgens"]].sum()
            tabela.loc[i, "TOTAL"] = soma_TOTAL
    
    #===== Tx. Elegível =====#
    for i in range(len(tabela["Tx. Elegível"])):
        denominador = ( tabela.loc[i, "TOTAL"] - tabela.loc[i, "Virgens"] )
        numerador = tabela.loc[i, "Total Elegíveis"]
        if denominador == 0:
            tabela.loc[i, "Tx. Elegível"] = 0
        else:
            tabela.loc[i, "Tx. Elegível"] = numerador / denominador

    #===== Tx. Sucesso =====#
    for i in range(len(tabela["Tx. Sucesso"])):
        denominador = tabela.loc[i, "Total Elegíveis"]
        numerador = tabela.loc[i, "20"]
        if denominador == 0:
            tabela.loc[i, "Tx. Sucesso"] = 0
        else:
            tabela.loc[i, "Tx. Sucesso"] = numerador / denominador
    
    #===== Recusa Agenda | Recusa Entrev. | Recusa Total =====#
    for i in range(len(tabela["Recusa Agenda"])):
        denominador = tabela.loc[i, "Total Elegíveis"]
        numerador_rec_agenda = tabela.loc[i, "4"]
        numerador_rec_entrev = tabela.loc[i, "44"]
        if denominador == 0:
            tabela.loc[i, "Recusa Agenda"] = 0
            tabela.loc[i, "Recusa Entrev."] = 0
        else:
            tabela.loc[i, "Recusa Agenda"] = numerador_rec_agenda / denominador
            tabela.loc[i, "Recusa Entrev."] = numerador_rec_entrev / denominador
        tabela.loc[i, "Recusa Total"] = (tabela.loc[i, "Recusa Agenda"] + tabela.loc[i, "Recusa Entrev."])
    
    #===== Tratar a última linha de sub Total e rodar todas as formulas para calcular as taxas novamente =====#
    tabela.iloc[ (int(REPLICAS[0])), : ] = tabela.iloc[ :(int(REPLICAS[0])), : ].sum()

    for i in range(len(tabela["Tx. Elegível"])):
        denominador = ( tabela.loc[i, "TOTAL"] - tabela.loc[i, "Virgens"] )
        numerador = tabela.loc[i, "Total Elegíveis"]
        if denominador == 0:
            tabela.loc[i, "Tx. Elegível"] = 0
        else:
            tabela.loc[i, "Tx. Elegível"] = numerador / denominador


    for i in range(len(tabela["Tx. Sucesso"])):
        denominador = tabela.loc[i, "Total Elegíveis"]
        numerador = tabela.loc[i, "20"]
        if denominador == 0:
            tabela.loc[i, "Tx. Sucesso"] = 0
        else:
            tabela.loc[i, "Tx. Sucesso"] = numerador / denominador


    for i in range(len(tabela["Recusa Agenda"])):
        denominador = tabela.loc[i, "Total Elegíveis"]
        numerador_rec_agenda = tabela.loc[i, "4"]
        numerador_rec_entrev = tabela.loc[i, "44"]
        if denominador == 0:
            tabela.loc[i, "Recusa Agenda"] = 0
            tabela.loc[i, "Recusa Entrev."] = 0
        else:
            tabela.loc[i, "Recusa Agenda"] = numerador_rec_agenda / denominador
            tabela.loc[i, "Recusa Entrev."] = numerador_rec_entrev / denominador
        tabela.loc[i, "Recusa Total"] = (tabela.loc[i, "Recusa Agenda"] + tabela.loc[i, "Recusa Entrev."])
    
    tabela.iloc[(int(REPLICAS[j])), 0] = "TOTAL"
    
    tabela.columns = pd.MultiIndex.from_product([[cidade], tabela.columns])
    
    bd_todas_cidades.append(tabela)
    
    print(bd_todas_cidades[j])



#=== Salvar em uma planilha em excel ===#
# Crie uma nova planilha Excel
wb = Workbook()
ws = wb.active
ws.title = "Relatorio_Vigitel"

# Define o estilo de preenchimento para o fundo do cabeçalho
header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Define o alinhamento centralizado para o cabeçalho
header_alignment = Alignment(horizontal="center", vertical="center")

# Loop pelos DataFrames e escrevendo na planilha
row_offset = 1  # Inicializa a contagem de linhas na planilha
for i, df in enumerate(bd_todas_cidades):
    # Converter o DataFrame para linhas que o openpyxl pode usar
    rows = dataframe_to_rows(df, index=False, header=True)
    
    # Escreve os dados no Excel
    for j, row in enumerate(rows):
        ws.append(row)

        # Estiliza apenas o cabeçalho
        if j ==0 or j == 1:  # Cabeçalhos do multiíndice
            for col in range(1, len(row) + 1):  # Itera sobre as colunas
                cell = ws.cell(row=row_offset + j, column=col)
                cell.fill = header_fill
                cell.alignment = header_alignment

        # Formata como porcentagem colunas de 24 a 28
        if j > 1:  # Dados (exclui os cabeçalhos)
            for col in range(24, 29):  # Colunas 24 a 28
                cell = ws.cell(row=row_offset + j, column=col)
                if isinstance(cell.value, (int, float)):  # Verifica se é número
                    cell.number_format = '0.0%'  # Formato de porcentagem com uma casa decimal
                    cell.value = float(cell.value)  # Converte o valor para floats
    
    # Adicionar 1 linha em branco entre os DataFrames, exceto no último
    if i < (len(bd_todas_cidades) - 1):
        row_offset = ws.max_row + 2  # Ajusta o offset para o próximo DataFrame
        for _ in range(1):  # Adiciona linhas vazias
            ws.append([])
    else:
        row_offset = ws.max_row + 1

# Salvar o arquivo Excel
wb.save("Relatorio_Vigitel_Fixo.xlsx")





# #=== Salvar em uma planilha em excel ===#
# # Crie uma nova planilha Excel
# wb = Workbook()
# ws = wb.active
# ws.title = "Relatorio_Vigitel"

# # Define o estilo de preenchimento para o fundo do cabeçalho
# header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# # Loop pelos DataFrames e escrevendo na planilha
# row_offset = 1  # Inicializa a contagem de linhas na planilha
# for i, df in enumerate(bd_todas_cidades):
#     # Converter o DataFrame para linhas que o openpyxl pode usar
#     rows = dataframe_to_rows(df, index=False, header=True)
    
#     # Escreve os dados no Excel
#     for j, row in enumerate(rows):
#         ws.append(row)

#         # Estiliza apenas o cabeçalho
#         if j == 1:  # Cabeçalho da tabela (primeira linha do DataFrame)
#             for col in range(1, len(row) + 1):  # Itera sobre as colunas
#                 ws.cell(row=row_offset + j, column=col).fill = header_fill
    
#     # Adicionar 1 linha em branco entre os DataFrames, exceto no último
#     if i < (len(bd_todas_cidades) - 1):
#         row_offset = ws.max_row + 2  # Ajusta o offset para o próximo DataFrame
#         for _ in range(1):  # Adiciona linhas vazias
#             ws.append([])
#     else:
#         row_offset = ws.max_row + 1

# # Salvar o arquivo Excel
# wb.save("Relatorio_Vigitel_Fixo.xlsx")



# #=== Salvar em uma planilha em excel ===#
# # Crie uma nova planilha Excel
# wb = Workbook()
# ws = wb.active
# ws.title = "Relatorio_Vigitel"

# # Loop pelos DataFrames e escrevendo na planilha
# for i, df in enumerate(bd_todas_cidades):
#     # Converter o DataFrame para linhas que o openpyxl pode usar
#     rows = dataframe_to_rows(df, index=False, header=True)
    
#     # Escreve os dados no Excel
#     for row in rows:
#         ws.append(row)
    
#     # Adicionar 1 linha em branco entre os DataFrames, exceto no último
#     if i < len(bd_todas_cidades) - 1:
#         ws.append([])

# # Salvar o arquivo Excel
# wb.save("dataframes_cidades.xlsx")
